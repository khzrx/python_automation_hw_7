import csv
import constants as c
from pypdf import PdfReader
from openpyxl import load_workbook


class TestFilesContents:
    def test_pdf_file_content(self, zip_file):
        with zip_file.open(c.FILE_NAMES['pdf']) as pdf_file:
            reader = PdfReader(pdf_file)

            assert len(reader.pages) == 2
            assert reader.pages[0].extract_text().startswith('GIT CHEAT SHEET')
            assert reader.pages[-1].extract_text().endswith('matches or wildcard globs.')

    def test_csv_file_content(self, zip_file):
        with zip_file.open(c.FILE_NAMES['csv']) as csv_file:
            content = csv_file.read().decode('utf-8')
            reader = list(csv.reader(content.splitlines()))

            assert len(reader) == 84
            assert reader[0] == ['Machete', '2010', '72']
            assert reader[-1] == ['Men of Honor', '2000', '41']

    def test_xlsx_file_content(self, zip_file):
        with zip_file.open(c.FILE_NAMES['xlsx']) as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.title == 'Okizeme'
            assert sheet.cell(1, 1).value == 'Knockdown'
